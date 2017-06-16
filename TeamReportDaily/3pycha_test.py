#!/usr/bin/env python
# -*- coding: utf-8 -*-

import cairo
import pycha.pie
import pycha.bar
import pycha.scatter
import pycha.stackedbar
import pycha.line

import decimal
import operator
from openpyxl import load_workbook

#load_data_global_ver
personal_time_list = []
create_test_case_cost_list = []
execute_test_case_cost_list = []
ticket_cost_list = []
name_code = {}
#In order to make a fixed pair task_list:color, should hard code task tpye
task_list=["CreateTestCase", "ExecuteTestCase", "AutomateTestCase", "BugVerification", "TroubleShooting", "Meeting", "StandUp", "Documentation", "ExploratoryTesting", "Other-Automation", "CreateBugTicket", "Testing-Data"]


def load_data():
    wb = load_workbook('TeamTimeSheet.xlsx')
    sheet_ranges = wb['Sheet']
    max_line=sheet_ranges.max_row
    print max_line



    title_col={
        sheet_ranges["A1"].value : "A",
        sheet_ranges["B1"].value: "B",
        sheet_ranges["C1"].value: "C",
        sheet_ranges["D1"].value: "D",
        sheet_ranges["E1"].value: "E",
        sheet_ranges["F1"].value: "F",
        sheet_ranges["G1"].value: "G",
        sheet_ranges["H1"].value: "H"
        }

    tester_col = title_col["Tester"]
    team_col = title_col["Team"]
    test_case_id_col = title_col["Test Case Id"]
    ticket_id_col = title_col["Ticket Id"]
    duration_col = title_col["Duration"]
    start_date_col = title_col["StartDate"]
    end_date_col = title_col["EndDate"]
    task_col = title_col["Task"]




    big_dic = {}
    for line in range(2,max_line+1):
        #get_pos
        tester_pos = tester_col + str(line)
        team_pos = team_col + str(line)
        test_case_pos = test_case_id_col + str(line)
        ticket_id_pos = ticket_id_col + str(line)
        duration_pos = duration_col+ str(line)
        start_date_pos = start_date_col+ str(line)
        end_date_pos = end_date_col+ str(line)
        task_pos = task_col+ str(line)
        #get_value
        tester = sheet_ranges[tester_pos].value.split(' ')[0]
        team = sheet_ranges[team_pos].value
        test_case_id = sheet_ranges[test_case_pos].value
        ticket_id = sheet_ranges[ticket_id_pos].value
        duration = sheet_ranges[duration_pos].value
        start_date = sheet_ranges[start_date_pos].value
        end_date = sheet_ranges[end_date_pos].value
        task = sheet_ranges[task_pos].value

        big_dic[line] = str(tester) + "\t" + str(team) + "\t" + str(test_case_id) + "\t" + str(ticket_id) + "\t" + str(duration) + "\t" + str(start_date) + "\t" + str(end_date) + "\t" + str(task)

    #data for personal time distribution
    #name_code
    #key:name #value: code
    index = 0
    temp_personal_task_dic = {}
    temp_personal_create_test_case_cost_dic = {}
    temp_personal_execute_test_case_cost_dic = {}
    temp_personal_ticket_cost_dic = {}
    for line in big_dic:    
        values = big_dic[line].split("\t")
        tester = values[0]
        #print line
        #print "$" * 32
        duration = int(values[4])
        test_case_id = values[2]
        ticket_id = values[3]
        task = values[7]

        if tester not in name_code:
            name_code[tester] = index
            index =  index + 1
        #if task not in task_list:
         #   task_list.append(task)
        
        temp_key = tester+"\t"+task
        if temp_key not in temp_personal_task_dic:
            temp_personal_task_dic[temp_key]=duration
        else:
            temp_personal_task_dic[temp_key]+=duration
            
        temp_key = tester + "\t" + test_case_id
        if task == "CreateTestCase":
            if temp_key not in temp_personal_create_test_case_cost_dic:
                temp_personal_create_test_case_cost_dic[temp_key]=duration
            else:
                temp_personal_create_test_case_cost_dic[temp_key]+=duration
        if task == "ExecuteTestCase":
            if temp_key not in temp_personal_execute_test_case_cost_dic:
                temp_personal_execute_test_case_cost_dic[temp_key]=duration
            else:
                temp_personal_execute_test_case_cost_dic[temp_key]+=duration
        if ticket_id != 'None':    
            temp_key = tester + "\t" + ticket_id
            if temp_key not in temp_personal_ticket_cost_dic:
                temp_personal_ticket_cost_dic[temp_key]=duration
            else:
                temp_personal_ticket_cost_dic[temp_key]+=duration
    
        
    
    

        
    #personal time list
    #init personal_time_list
    for person_number in range(index):
        temp_dic = {}
        personal_time_list.append(temp_dic)
        
    for name_task, duration in temp_personal_task_dic.iteritems():
        #print name_task, duration
        keys = name_task.split("\t")
        name_record = keys[0]
        task_record = keys[1]
        code = name_code[name_record]
        dic = personal_time_list[code]
        dic[task_record] = duration



    #print personal_time_list

    #create test case
    for person_number in range(index):
        temp_dic = {}
        create_test_case_cost_list.append(temp_dic)

    for name_testCaseId, duration in temp_personal_create_test_case_cost_dic.iteritems():
        #print name_task, duration
        keys = name_testCaseId.split("\t")
        name_record = keys[0]
        testCaseId_record = keys[1]
        code = name_code[name_record]
        dic = create_test_case_cost_list[code]
        #if testCaseId_record != 'None':
        dic[testCaseId_record] = duration

    #print create_test_case_cost_list

    #execute test case

    for person_number in range(index):
        temp_dic = {}
        execute_test_case_cost_list.append(temp_dic)

    for name_testCaseId, duration in temp_personal_execute_test_case_cost_dic.iteritems():
        #print name_task, duration
        keys = name_testCaseId.split("\t")
        name_record = keys[0]
        testCaseId_record = keys[1]
        code = name_code[name_record]
        dic = execute_test_case_cost_list[code]
        #if testCaseId_record != 'None':
        dic[testCaseId_record] = duration
    #print execute_test_case_cost_list

    #ticket cost
    for person_number in range(index):
        temp_dic = {}
        ticket_cost_list.append(temp_dic)

    for name_ticketId, duration in temp_personal_ticket_cost_dic.iteritems():
        #print name_task, duration
        keys = name_ticketId.split("\t")
        name_record = keys[0]
        ticketId_record = keys[1]
        code = name_code[name_record]
        dic = ticket_cost_list[code]
        dic[ticketId_record] = duration
    #print ticket_cost_list

#paint
g_data_vec = []




options={   
        'titleFont':'字体',
        'title': '', #图片标题  
        'labelColor':'#666666',
        'labelFont':'字体',
        'labelFontSize':12,
        'labelWidth':50.0,
        'tickFont':'字体',
        'tickFontSize':9,
        'background': {  
            'chartColor': '#ffffff',     #图表背景色  
            'baseColor': '#ffffff',      #边框颜色  
            'lineColor': '#0000ff'       #横线颜色  
        },  
        'colorScheme': {  
            'name': 'gradient',
            'args': {
               'initialColor': '#8EC2F5', #图表颜色 
                },
           
        },  
        'legend': {  
            'hide': False,     #是否隐藏图标示例  
        },  
        'padding': {  
            'left': 80,       #左边框  
            'right':50,
            'top':50,
            'bottom': 50,     #底边框  
        },
        'axis': {
            'labelFontSize':12,
            'x':{
            'label':'ID',
            },
            'y':{
             'label': "Minutes",
             
            },
            },
     
        
}

names = (
    ('TingTing', 0),
    ('Leon', 1),
    ('Ella', 2),
    ('Sowmya', 3),
    ('Deepthi', 4),
    ('Tracis', 5),
    ('Wilson', 6)
)


colors=['#FF3349','#F633FF','#9A33FF', '#335AFF','#33FFB3', '#FFF233', '#FF9C33','#8EC2F5','#D5FFC0', '#33FFFD', '#EEEEEE', '#6D9D88', '82D7FE']

teamOptions={   
         'titleFont':'字体',
        'title': '', #图片标题  
        'labelColor':'#666666',
        'labelFont':'字体',
        'labelFontSize':9,
        'labelWidth':50.0,
        'tickFont':'字体',
        'tickFontSize':9,
        'background': {  
            'chartColor': '#ffffff',     #图表背景色  
            'baseColor': '#ffffff',      #边框颜色  
            'lineColor': '#0000ff'       #横线颜色  
        },  
        'colorScheme': {  
            'name': 'gradient',
            'args': {
               'initialColor': '#8EC2F5', #图表颜色 
                },
 
        },  
        'legend': {  
            'hide': False,     #是否隐藏图标示例
            'position':{
                'top':0,
                'left':0,
                }
        },  
        'padding': {  
            'left': 50,       #左边框  
            'right':50,
            'top':50,
            'bottom': 50,     #底边框  
        },
        'axis': {
            'labelFontSize':12,
        'x':{
            'ticks':[dict(v=i, label=l[0]) for i,l in enumerate(names)],
            'label':'ID',
            'rotate': 0,
            'tickCount': 10,
            },
        'y':{
             'label': "Minutes",
             
            },
            },
        
}

teamOptions_custom = teamOptions
for key, value in teamOptions.iteritems():
    teamOptions_custom[key] = value
    

#设置画布
def set_charvalue():
    width,height=600,600 
    surface=cairo.ImageSurface(cairo.FORMAT_ARGB32,width,height) 
    return surface
    
#画饼图
def draw_pie(surface, options, dataSet, file_name):
    chart=pycha.pie.PieChart(surface,options) 
    chart.addDataset(dataSet) 
    chart.render() 
    surface.write_to_png(file_name) 

#垂直直方图
def draw_vertical_bar(surface, options, dataSet, file_name):
    chart=pycha.bar.VerticalBarChart(surface,options) 
    chart.addDataset(dataSet) 
    chart.render() 
    surface.write_to_png(file_name)   
 
#垂直水平直方图    
def draw_horizontal_bar(surface, options, dataSet, file_name):
    chart = pycha.bar.HorizontalBarChart(surface,options) 
    chart.addDataset(dataSet) 
    chart.render() 
    surface.write_to_png(file_name)   
    
#线图    
def draw_line(surface, options, dataSet,file_name):
    chart = pycha.line.LineChart(surface,options) 
    chart.addDataset(dataSet) 
    chart.render() 
    surface.write_to_png(file_name)      

#点图    
def draw_scatterplot(surface, options, dataSet, file_name):
    chart = pycha.scatter.ScatterplotChart(surface,options) 
    chart.addDataset(dataSet) 
    chart.render() 
    surface.write_to_png(file_name)         

#垂直块图     
def draw_stackedverticalbarChar(surface, options, dataSet, file_name):
    chart = pycha.stackedbar.StackedVerticalBarChart(surface,options) 
    chart.addDataset(dataSet) 
    chart.render() 
    surface.write_to_png(file_name)      

#水平块图
def draw_stackedhorizontalbarChart(surface, options, dataSet, file_name):
    chart = pycha.stackedbar.StackedHorizontalBarChart(surface,options) 
    chart.addDataset(dataSet) 
    chart.render() 
    surface.write_to_png(file_name)    
    
     
    
    
def drawPersonalTime(taskType,name, dic):

    title  = name + "\'s " + taskType + 'Time Cost'

    file_name = 'img/' +name + taskType  + 'TimeCost.png'
    file_name = file_name.replace(' ','')
    dataSet = ()
    id_list = []
    index = 0
    id_tuple_list = []
    for key,value in dic.iteritems():
        id_list.append([index,value])
        temp_id_pair = (key, index)
        id_tuple_list.append(temp_id_pair)
        index = index + 1
    id_tuple = tuple(id_tuple_list)
    dataSet_list = [(title, id_list)]
    if id_list == []:
        return
    dataSet = tuple(dataSet_list)
    #print dataSet
    #print "#" * 32
        
    
    #图像属性定义
    teamOptions['title'] = title
    teamOptions['axis']['x']['ticks']=[dict(v=i, label=l[0]) for i,l in enumerate(id_tuple)]
    teamOptions['axis']['x']['tickCount'] = len(id_tuple)
    if len(id_tuple) > 10:
        teamOptions['axis']['x']['rotate'] = 90
    teamOptions['legend']['hide'] = True
   
    surface = set_charvalue()
    
    draw_vertical_bar(surface, teamOptions, dataSet,file_name)
    if len(id_tuple) > 10:
        teamOptions['axis']['x']['rotate'] = 0


def drawTimeDistribution(name, dic):
    file_name = 'img/'+ name +'TimeDistribute.png'
    dataSet = ()
    dataSet_list = []
    for key, value in dic.iteritems():
        temp_tuple = ()
        temp_tuple_list = []
        temp_list = []
        temp_sub = [0,]
        temp_sub.append(int(value))
        temp_list.append(temp_sub)
        temp_tuple_list.append(key)
        temp_tuple_list.append(temp_list)
        temp_tuple = tuple(temp_tuple_list)
        dataSet_list.append(temp_tuple)
    dataSet = tuple(dataSet_list)
    print dataSet
    print '*' * 32
    options['title'] = name + '\'s Time Distribution'
    surface = set_charvalue()
    options['legend']['hide'] = True
    options['pieRadius'] = 0.4
    draw_pie(surface, options, dataSet,file_name)

def drawTeamTimeDistribution(name_code, personal_time_list):
        file_name = 'img/TeamTimeDistribution.png'
        dataSet = ()
        dataSet_list = []
        fake_list=[]

        for i in range(len(name_code)):
            fake_list.append([i,0.1])
        
        
        for task in task_list:
            temp_tuple = ()
            temp_tuple_list = []
            temp_list = []
            for name,code in name_code.iteritems():
                temp_sub_list = []
                temp_sub_list.append(int(code))
                dic = personal_time_list[int(code)]
                if task not in dic:
                    task_cost = 0
                else:
                    task_cost = dic[task]
                temp_sub_list.append(task_cost)
                temp_list.append(temp_sub_list)
            temp_tuple_list.append(task)
            temp_tuple_list.append(temp_list)
            temp_tuple = tuple(temp_tuple_list)
            dataSet_list.append(temp_tuple)
        dataSet_list.append(('Placeholder', fake_list))
        #order dataSet
        
        dataSet=tuple(dataSet_list)
        print dataSet
        print '#' * 32
        

        
        teamOptions['title'] = 'QA Team Time Distribution '
        names_tuple = ()
        name_tuple_list = []


        order_name_code_tuple=()
        sorted_name_code_list = sorted(name_code.items(), key=operator.itemgetter(1))

        #print sorted_name_code_list
        order_name_code_tuple=tuple(sorted_name_code_list)
        #print order_name_code_tuple
        #print '++'* 32
            

        
        teamOptions_custom['axis']['x']['ticks']=[dict(v=i, label=l[0]) for i,l in enumerate(order_name_code_tuple)]
        teamOptions_custom['legend']['position']['top'] = 10
        teamOptions_custom['legend']['position']['left'] = 10
        teamOptions_custom['legend']['hide'] = False
        teamOptions_custom['padding']={  
            'left': 120,       #左边框  
            'right':10,
            'top':100,
            'bottom': 10,     #底边框  
        }
        color_scheme_dic = teamOptions_custom['colorScheme']
        data_set_num = len(dataSet)
        print data_set_num
        colors_used = colors[:data_set_num]
        print colors_used
        color_scheme_dic_new = {
            'name': 'fixed',
            'args': {
                'colors' : colors_used,
                }
            }
        teamOptions_custom['colorScheme'] = color_scheme_dic_new
        surface = set_charvalue()

        
        draw_stackedverticalbarChar(surface, teamOptions_custom, dataSet, file_name)
        
    
    
    
    
if __name__ == '__main__':

    load_data()
    for name, code in name_code.iteritems():
        dic = personal_time_list[code]
        drawTimeDistribution(name, dic)
    
    for name, code in name_code.iteritems():
        dic = create_test_case_cost_list[code]
        drawPersonalTime('Create Test Case ', name, dic)
    for name, code in name_code.iteritems():
        dic = execute_test_case_cost_list[code]
        drawPersonalTime('Execute Test Case ', name, dic)
    for name, code in name_code.iteritems():
        dic = ticket_cost_list[code]
        drawPersonalTime('Ticket ', name, dic)
    drawTeamTimeDistribution(name_code, personal_time_list)
        
   
    
    
    
  
    
 
