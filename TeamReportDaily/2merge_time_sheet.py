#excel
from openpyxl import Workbook
from openpyxl import load_workbook
from os import listdir
#email
import smtplib
import sys
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage


#execl
dest_wb = Workbook()

dest_file = 'TeamTimeSheet.xlsx'
dest_ws = dest_wb.active

#first line
dest_ws["A1"] = "Tester"
dest_ws["B1"] = "Team"
dest_ws["C1"] = "Test Case Id"
dest_ws["D1"] = "Ticket Id"
dest_ws["E1"] = "Duration"
dest_ws["F1"] = "StartDate"
dest_ws["G1"] = "EndDate"
dest_ws["H1"] = "Task"
dest_ws["I1"] = "URL"
dest_ws["J1"] = "Note"


ws_cols = ["A","B","C","D", "E", "F","G","H","I","J"]

source_list = listdir("input")
des_line = 2
#email
sender = 'ella.sun@plexure.com'
subject = 'QA Team TimesSheet Check Result'
smtpserver = 'smtp.office365.com'
username = sys.argv[1]
password = sys.argv[2]
receiver_str = sys.argv[3]
receiver = receiver_str.split(",")
strTM = sys.argv[4]
msgRoot = MIMEMultipart('related')
msgRoot['Subject'] = subject  +'(' + strTM + ')'

title = '<p><font size="6" face="Microsoft YaHei" align="center">Check Result('+strTM+'):</font></p><br><br>'
sub_str0 = ''

team_list = ["SDK Team", "SNT Team", "Pukeko Team", "Data Team", "QA Team", "OPS Team"]
task_list = ["CreateTestCase", "ExecuteTestCase", "AutomateTestCase", "BugVerification", "TroubleShooting", "Meeting", "StandUp", "Documentation", "ExploratoryTesting", "Other-Automation", "CreateBugTicket", "Testing-Data"]
#log error

for input_file in source_list:
        source_wb = load_workbook("input/"+input_file)
        person_data = '<p><font size="4" face="Microsoft YaHei"><li>'+input_file+'</font></p></li>'
        person_err = ""
        sheet_ranges = source_wb['Sheet1']
        max_line=sheet_ranges.max_row
        print max_line
        # for every timesheet, copy every cell
        for line in range(2,max_line+1):
                #if tester is none, consider it is an empty line
                if sheet_ranges["A"+str(line)].value == None:
                        continue
                if sheet_ranges["E"+str(line)].value == None:
                        person_err += '<br><font color="red">\t[ERROR] line ' + str(line) + ':field [Duration] can not be null. </font><br>'
                        continue
                if str(sheet_ranges["B"+str(line)].value) not in team_list:
                        person_err += '<br><font color="red">\t[ERROR] line ' + str(line) + ': field [Team] '+str(sheet_ranges["B"+str(line)].value) +' is not in the list ' + str(team_list) + '.</font><br>'
                        continue
                
                if str(sheet_ranges["H"+str(line)].value) not in task_list:
                        person_err += '<br><font color="red">\t[ERROR] line ' + str(line) + ': field [Task type] ' + str(sheet_ranges["H"+str(line)].value) +' is not in the list ' + str(task_list) + '.</font><br>'
                        continue


                if str(sheet_ranges["E"+str(line)].value).isdigit() == False:
                        person_err += '<br><font color="red">\t[ERROR] line ' + lstr(line) + ': field [Duration] should be numberic. </font><br>'
                        continue
                if str(sheet_ranges["H"+str(line)].value).find("TestCase") > 0 and sheet_ranges["C"+str(line)].value == None:
                        person_err += '<br><font color="orange">\t[WARNING] line ' + str(line) + ': field [Test case]  should have value since the task type is ' + str(sheet_ranges["H"+str(line)].value) + '  . </font><br>'
                if str(sheet_ranges["H"+str(line)].value).find("Bug") > 0 and sheet_ranges["D"+str(line)].value == None:
                        person_err += '<br><font color="orange">\t[WARNING] line ' + str(line) + ': field [Ticket Id]  should have value since the task type is ' + str(sheet_ranges["H"+str(line)].value) +'. </font><br>'
                if str(sheet_ranges["A"+str(line)].value).find(" ") > 0:
                        person_err += '<br>\t[NOTICE] line ' + str(line) + ': family name will not be recorded. <br>'
                for col_index in ws_cols:
                        field = sheet_ranges[col_index+str(line)].value
                        if field == None:
                                dest_ws[col_index+str(des_line)]  = ""
                        else:
                                dest_ws[col_index+str(des_line)] = str(field)
                des_line = des_line + 1

        sub_str0 += person_data + person_err + "<br>"
 







dest_wb.save(dest_file)

sub_str1 = '<p><font size="3" face="Microsoft YaHei">Regards</font></p><p><font size="3" face="Microsoft YaHei">Ella</font></p><br><br>'

contentStr = title + sub_str0 + sub_str1
msgText = MIMEText(contentStr,'html','utf-8')


msgRoot.attach(msgText)
smtp = smtplib.SMTP(smtpserver, 587)
smtp.starttls()
smtp.login(username, password)
smtp.sendmail(sender, receiver, msgRoot.as_string())
smtp.quit()




